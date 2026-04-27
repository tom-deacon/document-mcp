"""Excel/XLSX parser — ingests workbooks as sheet-level structured elements."""

import logging
from pathlib import Path
from typing import Any, Dict, List

logger = logging.getLogger(__name__)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available — .xlsx/.xls files cannot be parsed. "
                   "Install with: pip install openpyxl")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _col_letter(n: int) -> str:
    """Convert 1-based column number to Excel column letter (1->A, 27->AA)."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result or "A"


def _is_numeric(s: str) -> bool:
    try:
        float(s.strip().replace(",", "").replace("%", "").replace("$", "").replace("£", ""))
        return True
    except ValueError:
        return False


def _has_header_row(rows: List[List[str]]) -> bool:
    """Heuristic: first row is a header if it's mostly non-numeric text."""
    if len(rows) < 2:
        return False
    non_empty = [c for c in rows[0] if c.strip()]
    if not non_empty:
        return False
    numeric = sum(1 for c in non_empty if _is_numeric(c))
    return (1 - numeric / len(non_empty)) >= 0.5


def _rows_to_markdown(rows: List[List[str]], max_cols: int = 26) -> str:
    """Render a 2-D list of strings as a Markdown table, capped at max_cols."""
    if not rows:
        return ""
    ncols = min(max(len(r) for r in rows), max_cols)
    padded = [(r + [""] * ncols)[:ncols] for r in rows]

    def esc(s: str) -> str:
        return s.replace("|", "\\|").replace("\n", " ").strip()

    lines = ["| " + " | ".join(esc(c) for c in padded[0]) + " |"]
    lines.append("| " + " | ".join(["---"] * ncols) + " |")
    for row in padded[1:]:
        lines.append("| " + " | ".join(esc(c) for c in row) + " |")

    actual_cols = max(len(r) for r in rows)
    if actual_cols > max_cols:
        lines.append(f"*Truncated: showing {max_cols} of {actual_cols} columns*")

    return "\n".join(lines)


def _sheet_summary(name: str, rows: List[List[str]], has_header: bool) -> str:
    """One-line human description of a sheet."""
    if not rows:
        return f"Empty sheet: {name}"
    nrows = len(rows)
    ncols = max(len(r) for r in rows) if rows else 0
    if has_header and rows:
        headers = [c for c in rows[0] if c.strip()][:6]
        col_desc = ", ".join(headers)
        if len([c for c in rows[0] if c.strip()]) > 6:
            col_desc += " ..."
        return f"Sheet '{name}': {nrows - 1} data rows, columns: {col_desc}"
    return f"Sheet '{name}': {nrows} rows × {ncols} columns"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_excel_structured(file_path: Path) -> List[Dict[str, Any]]:
    """
    Parse an Excel workbook into one element per non-empty worksheet.

    Each element dict contains:
      element_type   — "sheet"
      text           — plain-text representation (sheet name + row data)
      section_path   — sheet name
      page_or_sheet  — sheet name
      html_or_markdown — Markdown table
      source_range   — e.g. "A1:G42"
      summary        — one-line description of the sheet content
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel parsing. "
            "Install with: pip install openpyxl"
        )

    logger.info("[EXCEL] Parsing: %s", file_path.name)

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open Excel file '{file_path.name}': {exc}") from exc

    elements: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Read all cell values, converting None -> ""
        all_rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(v) if v is not None else "" for v in row])

        # Trim trailing empty rows
        while all_rows and not any(c.strip() for c in all_rows[-1]):
            all_rows.pop()

        # Trim trailing empty columns
        if all_rows:
            max_used_col = max(
                max((i + 1 for i, c in enumerate(row) if c.strip()), default=0)
                for row in all_rows
            )
            all_rows = [row[:max_used_col] for row in all_rows]

        if not all_rows:
            logger.info("[EXCEL] Sheet '%s' is empty, skipping", sheet_name)
            continue

        has_header = _has_header_row(all_rows)
        summary = _sheet_summary(sheet_name, all_rows, has_header)
        md = _rows_to_markdown(all_rows)
        range_addr = f"A1:{_col_letter(max(len(r) for r in all_rows))}{len(all_rows)}"

        plain_lines = [f"Sheet: {sheet_name}", summary, ""]
        plain_lines += [" | ".join(row) for row in all_rows]
        plain_text = "\n".join(plain_lines)

        logger.info(
            "[EXCEL] Sheet '%s': %d rows × %d cols (header=%s)",
            sheet_name, len(all_rows),
            max(len(r) for r in all_rows) if all_rows else 0,
            has_header,
        )

        elements.append({
            "element_type": "sheet",
            "text": plain_text,
            "section_path": sheet_name,
            "page_or_sheet": sheet_name,
            "html_or_markdown": md,
            "source_range": range_addr,
            "summary": summary,
        })

    wb.close()
    logger.info("[EXCEL] %s: %d sheet(s) parsed", file_path.name, len(elements))
    return elements
